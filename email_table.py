import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# Step 1: Create the DataFrame
data = {
    'Jan-24': [1, 2, 3, 2, 1, 2],
    'Feb-24': [1, 2, 3, 2, 2, 1],
    'Mar-24': [1, 2, 1, 1, 1, 1],
    'Apr-24': [0, 1, 1, 0, 0, 1],
    'May-24': [1, 1, 1, 1, 1, 1],
    'Jun-24': [1, 2, 2, 1, 1, 1],
    'Jul-24': [2, 2, 1, 2, 2, 2],
    'Aug-24': [0, 0, 3, 1, 2, 2],
    'Sep-24': [0, 2, 2, 1, 1, 2]
}

rows = [
    ('America', 'CTO'),
    ('America', 'HKQ'),
    ('Asia', 'HKGX'),
    ('Asia', 'HKGX'),
    ('Asia', 'CPBG'),
    ('Asia', 'TSE')
]

# Create a DataFrame
df = pd.DataFrame(data, index=pd.MultiIndex.from_tuples(rows, names=['Region', 'City']))

# Step 2: Calculate row-wise totals (add a Total column for each row)
df['Total'] = df.sum(axis=1)

# Step 3: Calculate subtotals for each region (exclude 'Total' column when calculating by month)
region_totals = df.drop(columns='Total').groupby(level=0).sum()

# Calculate the total for the 'Total' column and append it to the region totals
region_totals['Total'] = df.groupby(level=0)['Total'].sum()

# Iterate through region_totals and append 'Region Total' rows
for region, total_row in region_totals.iterrows():
    # Create a new row with all the values, but for the index, add 'Total' to the region
    new_index = (region + ' Total', '')  # Region Total has an empty City ('')
    
    # Append the new row to the DataFrame
    df.loc[new_index] = total_row

# Step 4: Calculate the grand total
grand_total = df.sum(numeric_only=True)
# df.loc[('Grand Total', ''), :] = grand_total
df.loc[('Grand Total', '')] = grand_total

# Sort the DataFrame to keep the natural order (regions and totals together)
df = df.sort_index(level=0, sort_remaining=False)

# Step 5: Write to Excel using openpyxl
wb = Workbook()
ws = wb.active
ws.title = "Styled Report"

# Add the data from DataFrame to Excel
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

# Step 6: Style the Excel sheet
# Define styles: fills, fonts, borders, alignment
header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
blue_fill = PatternFill(start_color="E7F3FA", end_color="E7F3FA", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center_alignment = Alignment(horizontal="center", vertical="center")
bold_font = Font(bold=True)

# Style the header row
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = center_alignment
    cell.border = border

# Alternate row colors and style subtotal rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    region_cell = row[0]
    if "Total" in region_cell.value:
        # Apply gray fill for total rows
        for cell in row:
            cell.fill = gray_fill
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = border
    else:
        # Apply alternating blue fill for non-total rows
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = blue_fill
            cell.alignment = center_alignment
            cell.border = border

# Save the Excel file
excel_file = "styled_report.xlsx"
wb.save(excel_file)

print(f"Excel file '{excel_file}' created successfully.")